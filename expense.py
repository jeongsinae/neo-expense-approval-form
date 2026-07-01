"""
Expense Report Automation

This script demonstrates how to automatically generate an expenditure report
("지출결의서") from a set of credit‑card or cash receipts. It does the following:

* Performs OCR on each uploaded receipt PDF (using PyMuPDF + Tesseract) to
  extract key fields such as the transaction date, merchant name and total
  amount.
* Populates a provided Excel template with the extracted details using the
  openpyxl library.
* Merges all original receipt PDFs into a single PDF file for final proof
  (증빙자료) using the pypdf library.

Before running this example, install the required packages:

    pip install openpyxl pypdf PyMuPDF pillow pytesseract

You must also install Tesseract OCR and the Korean language data (`kor`)
pack. On most Linux systems, you can install these with your package manager
(e.g. `sudo apt install tesseract-ocr tesseract-ocr-kor`). See
https://tesseract-ocr.github.io/tessdoc/Data-Files-in-different-versions.html
for supported languages.

Note: This code is an example. You need to adapt the cell positions to match
your own Excel form. In a real application, consider adding a web interface
(Flask, FastAPI or Django) to handle file uploads, preview OCR results and
confirm before generating the final report.
"""

import re
import sys
from pathlib import Path
from typing import List, Dict

import fitz  # PyMuPDF
import pytesseract
from pypdf import PdfWriter
from openpyxl import load_workbook
from PIL import Image


# -----------------------------------------------------------------------------
# OCR and data extraction
# -----------------------------------------------------------------------------

def extract_text_from_pdf(pdf_path: Path) -> str:
    """Extract text from a PDF by rasterizing each page and running Tesseract.

    Args:
        pdf_path: Path to the receipt PDF.

    Returns:
        A string containing all OCR’d text from the document.
    """
    text = []
    with fitz.open(pdf_path) as doc:
        for page in doc:
            pix = page.get_pixmap(dpi=300)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            # Use both Korean and English language data when decoding receipts.
            ocr_result = pytesseract.image_to_string(img, lang="kor+eng")
            text.append(ocr_result)
    return "\n".join(text)


def parse_receipt(text: str) -> Dict[str, str]:
    """Parse a receipt’s OCR’d text to extract date, vendor and amount.

    This function uses regular expressions tailored to typical Korean credit‑card
    receipts. You should adjust the patterns to suit your receipts. If a field
    isn’t found, it returns an empty string.

    Args:
        text: OCR output from the receipt PDF.

    Returns:
        A dictionary with keys: 'date', 'vendor', 'amount', 'memo'.
    """
    # Normalise whitespace
    clean = re.sub(r"\s+", " ", text)

    # Date: match patterns like 2026-05-28, 2026/05/28, 2026.05.28
    date_match = re.search(r"(20\d{2})[-./](\d{1,2})[-./](\d{1,2})", clean)
    date = "" if date_match is None else f"{date_match.group(1)}-{int(date_match.group(2)):02d}-{int(date_match.group(3)):02d}"

    # Amount: only trust values after the target total labels.
    amount_match = re.search(
        r"(?:총\s*결제\s*원화\s*금액|합\s*계)\D{0,20}(\d{1,3}(?:[,.]\d{3})+|\d{4,9})",
        clean,
    )
    amount = "" if amount_match is None else amount_match.group(1).replace(",", "")

    # Vendor: simplistic – take the first 10 characters before the word '승인'
    vendor_match = re.search(r"([가-힣A-Za-z0-9\s]{2,20})\s*승인", clean)
    vendor = "" if vendor_match is None else vendor_match.group(1).strip()

    return {
        "date": date,
        "vendor": vendor,
        "amount": amount,
        "memo": "자동 추출" if date or amount or vendor else ""
    }


# -----------------------------------------------------------------------------
# Excel writing
# -----------------------------------------------------------------------------

def populate_excel(template_path: Path, records: List[Dict[str, str]], output_path: Path) -> None:
    """Fill the Excel template with receipt information.

    Args:
        template_path: Path to the Excel template.
        records: List of dictionaries with keys 'date', 'vendor', 'amount', 'memo'.
        output_path: Path to save the completed workbook.
    """
    wb = load_workbook(template_path)
    ws = wb.active

    # Example: start inserting data at row 7. Adjust cell references to your form.
    start_row = 7
    for idx, rec in enumerate(records):
        row = start_row + idx
        ws[f"B{row}"] = rec.get("date")
        ws[f"C{row}"] = rec.get("vendor")
        ws[f"D{row}"] = int(rec.get("amount", 0)) if rec.get("amount") else ""
        ws[f"E{row}"] = rec.get("memo")

    # Optionally compute total amount in D column and write to a designated cell.
    total = sum(int(r.get("amount", 0)) for r in records if r.get("amount"))
    ws["D20"] = total

    wb.save(output_path)


# -----------------------------------------------------------------------------
# PDF merging
# -----------------------------------------------------------------------------

def merge_pdfs(pdf_paths: List[Path], output_path: Path) -> None:
    """Combine multiple receipt PDFs into a single PDF.

    Args:
        pdf_paths: List of PDF file paths to merge.
        output_path: Destination path for the merged PDF.

    The pypdf documentation demonstrates that you can iterate through a list
    of PDF filenames, call `append` on a PdfWriter and then write the result
    to disk【345807601017852†L90-L101】.
    """
    writer = PdfWriter()
    for pdf in pdf_paths:
        writer.append(str(pdf))
    writer.write(str(output_path))


# -----------------------------------------------------------------------------
# High‑level function
# -----------------------------------------------------------------------------

def generate_expense_report(template_path: Path, receipt_pdf_paths: List[Path], excel_out: Path, pdf_out: Path) -> None:
    """High‑level convenience function to create an expense report.

    Args:
        template_path: Excel template path.
        receipt_pdf_paths: List of receipt PDF files.
        excel_out: Where to save the completed Excel file.
        pdf_out: Where to save the merged proof PDF.
    """
    records = []
    for pdf in receipt_pdf_paths:
        text = extract_text_from_pdf(pdf)
        record = parse_receipt(text)
        records.append(record)

    populate_excel(template_path, records, excel_out)
    merge_pdfs(receipt_pdf_paths, pdf_out)


# -----------------------------------------------------------------------------
# Command-line interface
# -----------------------------------------------------------------------------

def main(args=None):
    import argparse

    parser = argparse.ArgumentParser(description="Generate an expense report from receipts.")
    parser.add_argument("template", type=str, help="Path to Excel template (.xlsx)")
    parser.add_argument("receipts", type=str, nargs="+", help="Paths to receipt PDFs")
    parser.add_argument("--excel-out", type=str, default="expense_report.xlsx", help="Output Excel filename")
    parser.add_argument("--pdf-out", type=str, default="receipts_merged.pdf", help="Output merged PDF filename")
    args = parser.parse_args(args)

    template_path = Path(args.template)
    receipts = [Path(p) for p in args.receipts]
    excel_out = Path(args.excel_out)
    pdf_out = Path(args.pdf_out)

    generate_expense_report(template_path, receipts, excel_out, pdf_out)
    print(f"Excel report saved to {excel_out}\nMerged PDF saved to {pdf_out}")


if __name__ == "__main__":
    main()
