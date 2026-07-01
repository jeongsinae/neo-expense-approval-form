"""
지출결의서 자동 생성 웹앱.

- 영수증 이미지/PDF를 업로드하면 OCR로 금액을 뽑아내고
- 4월 템플릿을 복사해 지출결의서 Excel을 만들고
- 영수증을 하나의 증빙 PDF로 병합해 다운로드 링크를 돌려준다.
"""

from __future__ import annotations

import io
import os
import re
import shutil
import subprocess
import sys
import uuid
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl
from flask import Flask, Response, jsonify, render_template_string, request, send_file, url_for
from openpyxl.utils import get_column_letter
from PIL import Image, ImageOps
from pypdf import PdfReader, PdfWriter


BASE_DIR = Path(__file__).resolve().parent
# 템플릿은 레포에 포함된 파일을 우선 사용하고, 없으면 로컬 참조 경로로 폴백한다.
_REPO_TEMPLATE = BASE_DIR / "template" / "template.xlsx"
_LOCAL_TEMPLATE = Path(r"C:\Users\ddoch\Desktop\지출결의서\4월\26년_04월_지출결의서_정시내.xlsx")
TEMPLATE_XLSX = _REPO_TEMPLATE if _REPO_TEMPLATE.exists() else _LOCAL_TEMPLATE
WORK_DIR = BASE_DIR / "work"
OUTPUT_DIR = BASE_DIR / "generated"
OCR_SCRIPT = BASE_DIR / "ocr_windows.ps1"
CI_DIR = BASE_DIR / "CI"

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".tif", ".tiff"}
PDF_EXTS = {".pdf"}
UPLOAD_EXTS = IMAGE_EXTS | PDF_EXTS
SKIP_STEMS = ("지출결의서", "지출증빙", "거래목록", "검토용", "정산서")

# 지출결의서 데이터 행 범위 (4월 템플릿 기준)
EXPENSE_SHEET_INDEX = 1   # 두 번째 시트: 첨부1.실행예산상세정산
DATA_START_ROW = 9
DATA_END_ROW = 76         # 최대 68건
TITLE_CELL = "B2"

# 계정 매핑 — 4월 템플릿 J/K/L 컬럼(계정번호/비목/세목) 전체 (68개)
ACCOUNTS: dict[int, str] = {
    # 복리후생비
    1: "조식", 2: "석식", 3: "휴일", 4: "간식", 5: "회식",
    # 접대비
    6: "경조", 7: "식대", 8: "선물",
    # 회의비
    9: "회의실사용료", 10: "회의식대", 11: "회의음료", 12: "물품구매",
    # 여비교통비
    13: "숙박", 14: "항공", 15: "여객선", 16: "철도", 17: "버스",
    18: "택시", 19: "지하철", 20: "자가차량",
    # 차량유지비
    21: "차량주유비", 22: "선박주유비", 23: "통행료", 24: "수리비",
    25: "주차비", 26: "차량렌탈료",
    # 통신비
    27: "인터넷", 28: "전화", 29: "우편",
    # 수도광열비
    30: "전력", 31: "가스수도",
    # 지급임차료
    32: "사무실임차료", 33: "선박용선료", 34: "기타임차료",
    # 도서인쇄비
    35: "도서구입", 36: "인쇄비",
    # 지급수수료
    37: "사무실관리비", 38: "성과심사비", 39: "보안경비", 40: "기타수수료",
    # 장비임차료
    41: "컴퓨터", 42: "복합기", 43: "사무용가구",
    44: "선박(외부)", 45: "측량장비(외부)",
    # 운영비
    46: "수선비", 47: "발송비", 48: "사무용품비", 49: "잡비",
    # 비통제/기타
    50: "제안비용", 51: "세금과공과금", 52: "보험료", 53: "보증금(이자비용)",
    # 자사 사용료
    54: "선박", 55: "측량장비", 56: "차량", 57: "사무실 임대", 58: "S/W",
    # 수당
    59: "야근/휴일수당", 60: "파견수당", 61: "현장/안전수당", 62: "PM수당",
    # 외주비
    63: "계약직", 64: "하도급",
    # 재료비
    65: "재료비",
    # 외주비 공통비 / 재료비 공통비
    66: "계약직 공통비", 67: "하도급 공통비", 68: "재료비 공통비",
}

# 비목 (대분류) — UI 드롭다운에서 그룹 표기용
ACCOUNT_CATEGORY: dict[int, str] = {
    **{n: "복리후생비" for n in range(1, 6)},
    **{n: "접대비" for n in range(6, 9)},
    **{n: "회의비" for n in range(9, 13)},
    **{n: "여비교통비" for n in range(13, 21)},
    **{n: "차량유지비" for n in range(21, 27)},
    **{n: "통신비" for n in range(27, 30)},
    **{n: "수도광열비" for n in range(30, 32)},
    **{n: "지급임차료" for n in range(32, 35)},
    **{n: "도서인쇄비" for n in range(35, 37)},
    **{n: "지급수수료" for n in range(37, 41)},
    **{n: "장비임차료" for n in range(41, 46)},
    **{n: "운영비" for n in range(46, 50)},
    **{n: "비통제/기타" for n in range(50, 54)},
    **{n: "자사 사용료" for n in range(54, 59)},
    **{n: "수당" for n in range(59, 63)},
    **{n: "외주비" for n in range(63, 65)},
    65: "재료비",
    **{n: "외주비 공통비" for n in range(66, 68)},
    68: "재료비 공통비",
}

# 자주 쓰는 계정의 기본 목적 문구
DEFAULT_PURPOSE: dict[int, str] = {
    10: "회의식대(중식)",
    29: "등기우편",
    36: "인쇄비",
    48: "사무용품",
    65: "정기결제비용",
}

# 가맹점명 → 계정번호 추론 (확실한 키워드만)
VENDOR_TO_ACCOUNT = (
    (("chatgpt", "챗지피티", "claude", "클로드", "cursor", "커서", "kiri", "엔진", "engine"), 65),
    (("office", "오피스", "문구", "사무용품"), 48),
    (("소프트웨어", "인쇄", "출력", "프린트"), 36),
    (("등기", "우편"), 29),
    (("주유", "기름", "셀프주유"), 21),
    (("주차장",), 25),
    (("택시",), 18),
    (("지하철", "철도", "ktx"), 16),
)


@dataclass
class Entry:
    date: str
    account_no: int
    account: str
    amount: int
    purpose: str
    attendee: str
    user: str
    vendor: str
    filename: str
    ocr_amount: int = 0
    ocr_text: str = ""


app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 300 * 1024 * 1024


# ─────────────────────────────────────────────────────────────────────────────
# 파일/파일명 유틸
# ─────────────────────────────────────────────────────────────────────────────

def ensure_dirs() -> None:
    WORK_DIR.mkdir(exist_ok=True)
    OUTPUT_DIR.mkdir(exist_ok=True)


def clean_name(name: str) -> str:
    name = Path(name).name
    return re.sub(r'[<>:"/\\|?*]', "_", name)


def parse_filename(stem: str) -> tuple[int | None, int | None, int | None, str]:
    """파일명 stem에서 (연, 월, 일, 가맹점)을 뽑아낸다. 다양한 패턴 지원."""
    # `0107_가맹점` (MMDD)
    match = re.match(r"^(\d{2})(\d{2})[_\-\s]+(.+)$", stem)
    if match:
        month = int(match.group(1))
        day = int(match.group(2))
        if 1 <= month <= 12 and 1 <= day <= 31:
            return None, month, day, match.group(3).strip(" _-") or stem
    # `2025-05-08_가맹점`, `20250508_가맹점`, `IMG_20250508`, `receipt 2025.05.08`
    for pattern in (
        re.compile(r"(20\d{2})[-_./]?(\d{2})[-_./]?(\d{2})"),
        re.compile(r"(\d{2})[-_./](\d{2})[-_./](\d{2})(?!\d)"),  # YY-MM-DD
    ):
        m = pattern.search(stem)
        if not m:
            continue
        try:
            year, month, day = int(m.group(1)), int(m.group(2)), int(m.group(3))
        except ValueError:
            continue
        if year < 100:
            year += 2000
        if 2020 <= year <= 2099 and 1 <= month <= 12 and 1 <= day <= 31:
            vendor = (stem[:m.start()] + stem[m.end():]).strip(" _-") or stem
            return year, month, day, vendor
    return None, None, None, stem


def receipt_key(name: str) -> str | None:
    """중복 영수증(같은 stem, 다른 확장자)을 묶을 키. 확장자만 맞으면 통과."""
    path = Path(name)
    if path.suffix.lower() not in UPLOAD_EXTS:
        return None
    if any(marker in path.stem for marker in SKIP_STEMS):
        return None
    _, month, day, vendor = parse_filename(path.stem)
    if month is not None:
        return f"{month:02d}{day:02d}_{re.sub(r'\\s+', '', vendor).casefold()}"
    return re.sub(r"\s+", "", path.stem).casefold()


def prefer_pdf(existing: str, candidate: str) -> bool:
    """기존이 이미지인데 새 파일이 PDF면 교체."""
    existing_ext = Path(existing).suffix.lower()
    candidate_ext = Path(candidate).suffix.lower()
    return existing_ext in IMAGE_EXTS and candidate_ext == ".pdf"


def infer_account(vendor: str) -> int:
    text = vendor.casefold()
    for keywords, account_no in VENDOR_TO_ACCOUNT:
        if any(k in text for k in keywords):
            return account_no
    return 10  # 기본: 회의식대


def build_entry(filename: str, year: int, month: int, attendee: str = "정시내", user: str = "정시내(개인)") -> Entry:
    stem = Path(filename).stem
    parsed_year, parsed_month, parsed_day, vendor = parse_filename(stem)
    if parsed_month is not None:
        date_year = parsed_year or year
        date_month, date_day = parsed_month, parsed_day
    else:
        today = datetime.now()
        date_year, date_month, date_day = year, month, today.day
        vendor = stem
    account_no = infer_account(vendor)
    purpose = DEFAULT_PURPOSE.get(account_no) or ACCOUNTS[account_no]
    if account_no == 65 and vendor:
        purpose = f"{vendor.upper()} {DEFAULT_PURPOSE[65]}"
    return Entry(
        date=f"{date_year:04d}.{date_month:02d}.{date_day:02d}",
        account_no=account_no,
        account=ACCOUNTS[account_no],
        amount=0,
        purpose=purpose,
        attendee=attendee or "정시내",
        user=user or "정시내(개인)",
        vendor=vendor,
        filename=filename,
    )


# ─────────────────────────────────────────────────────────────────────────────
# OCR — Windows 내장 OCR 엔진 (ocr_windows.ps1) 사용
# ─────────────────────────────────────────────────────────────────────────────

def run_ocr(path: Path) -> str:
    """OCR 텍스트 추출.

    Windows에서 ocr_windows.ps1이 있으면 내장 OCR 엔진을 쓰고,
    그 외 환경(Linux 배포 등)에서는 Tesseract로 대체한다.
    """
    if sys.platform.startswith("win") and OCR_SCRIPT.exists():
        return _run_ocr_windows(path)
    return _run_ocr_tesseract(path)


def _run_ocr_windows(path: Path) -> str:
    try:
        completed = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass",
             "-File", str(OCR_SCRIPT), "-Path", str(path)],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=60,
            check=False,
        )
    except subprocess.TimeoutExpired:
        return ""
    raw = completed.stdout
    for encoding in ("utf-8-sig", "cp949", "utf-16"):
        try:
            return raw.decode(encoding).strip()
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="ignore").strip()


def _run_ocr_tesseract(path: Path) -> str:
    """Tesseract 기반 OCR. PDF는 첫 페이지를 렌더링해 인식한다."""
    try:
        import pytesseract
    except ImportError:
        return ""

    images: list[Image.Image] = []
    try:
        if path.suffix.lower() == ".pdf":
            try:
                import fitz  # PyMuPDF
            except ImportError:
                return ""
            with fitz.open(str(path)) as doc:
                if doc.page_count < 1:
                    return ""
                page = doc.load_page(0)
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                images.append(Image.open(io.BytesIO(pix.tobytes("png"))))
        else:
            images.append(Image.open(path))

        texts = []
        for img in images:
            texts.append(pytesseract.image_to_string(img, lang="kor+eng"))
        return "\n".join(texts).strip()
    except Exception:
        return ""


def _normalize_amount(token: str) -> int | None:
    token = token.strip()
    # 소수점 .00 형태 처리
    if re.search(r"[.,]\d{1,2}$", token):
        base, decimals = re.split(r"[.,](?=\d{1,2}$)", token, maxsplit=1)
        if decimals != "00":
            return None
        token = base
    digits = re.sub(r"[.,]", "", token) if re.search(r"[.,]\d{3}", token) else re.sub(r"\D", "", token)
    if not digits:
        return None
    amount = int(digits)
    return amount if 1000 <= amount <= 5_000_000 else None


_NUMBER_RE = re.compile(r"\d{1,3}(?:[,.]\d{3})+(?:[,.]\d{2})?|\d{4,9}")
# 사업자번호/카드번호/승인번호 등 금액이 아닌 숫자의 앞 컨텍스트
_NON_AMOUNT_PREFIX_RE = re.compile(
    r"(?:사업자|가맹점|승인|카드|회원|전화|매장|점포|연락|tel|사뗘자|사룝자)\s*(?:번호|번트|번호[:：]?)?\s*[:：\-]?\s*$",
    flags=re.IGNORECASE,
)
# 하이픈 패턴 (사업자번호 443-01-03395, 카드번호 4937-51-** 등)
_HYPHEN_NUMBER_RE = re.compile(r"\d+\s*-\s*\d*\s*$")


def _is_amount_context(text: str, abs_start: int) -> bool:
    prefix = text[max(0, abs_start - 24): abs_start]
    if _NON_AMOUNT_PREFIX_RE.search(prefix):
        return False
    # 하이픈 그룹의 일부 (e.g. "443-01-03395"의 03395) 제외
    if _HYPHEN_NUMBER_RE.search(prefix):
        return False
    return True


_COMMA_AMOUNT_RE = re.compile(r"\d{1,3}(?:,\d{3})+")
_WON_AMOUNT_RE = re.compile(
    r"(\d{1,3}(?:[,.]\d{3})+(?:[,.]\d{2})?|\d{4,9})\s*[\[\(]?\s*(?:원|KRW)\s*[\]\)]?",
    flags=re.IGNORECASE,
)


def extract_amount(text: str) -> int:
    """OCR 결과에서 결제 금액 추출.
    천단위 콤마(예: 12,950)나 원/KRW 마커가 붙은 숫자만 후보로 인정.
    사업자번호·카드번호·전화번호 등 마커 없는 숫자는 제외. 후보 중 최댓값 채택.
    """
    if not text:
        return 0
    candidates: list[int] = []

    # 1) 천단위 콤마 형식 — 1,085,967 / 12,950
    for match in _COMMA_AMOUNT_RE.finditer(text):
        if not _is_amount_context(text, match.start()):
            continue
        digits = match.group(0).replace(",", "")
        try:
            amount = int(digits)
        except ValueError:
            continue
        if 1000 <= amount <= 5_000_000 and not (2020 <= amount <= 2099):
            candidates.append(amount)

    # 2) 원/KRW 마커가 붙은 숫자 — 8,300원 / 10.000 원 / 29.000 [KRW]
    for match in _WON_AMOUNT_RE.finditer(text):
        if not _is_amount_context(text, match.start(1)):
            continue
        amount = _normalize_amount(match.group(1))
        if amount and not (2020 <= amount <= 2099):
            candidates.append(amount)

    return max(candidates, default=0)


_DATE_PATTERNS = (
    # 2025-05-08, 2025/05/08, 2025.05.08, 2026년 5월 8일
    re.compile(r"(20\d{2})\s*[-./년]\s*(\d{1,2})\s*[-./월]\s*(\d{1,2})(?:\s*일)?"),
    # 2026 5 28 (공백 구분, 신한카드)
    re.compile(r"(?<!\d)(20\d{2})\s+(\d{1,2})\s+(\d{1,2})(?!\d)"),
    # 26.0522 (YY.MMDD 컴팩트, KB 해외카드 이용일시)
    re.compile(r"(?<![\d.])(\d{2})\.(\d{2})(\d{2})(?!\d)"),
    # 2605.15 (YYMM.DD 컴팩트)
    re.compile(r"(?<![\d.])(\d{2})(\d{2})\.(\d{1,2})(?!\d)"),
    # 25-05-21 (YY-MM-DD 구분자 명시)
    re.compile(r"(?<![\d/])(\d{2})\s*[-./]\s*(\d{1,2})\s*[-./]\s*(\d{1,2})(?![\d/])"),
)


def extract_date_from_text(text: str) -> tuple[int, int, int] | None:
    if not text:
        return None
    for pattern in _DATE_PATTERNS:
        for match in pattern.finditer(text):
            try:
                year = int(match.group(1))
                month = int(match.group(2))
                day_str = match.group(3)
                day = int(day_str)
            except ValueError:
                continue
            if year < 100:
                year += 2000
            # OCR이 두자릿수 일자를 "1 1"로 잘라먹은 경우 보강: 직후 자릿수 하나가 더 있으면 결합 시도
            if 1 <= day <= 3:
                after = text[match.end(): match.end() + 4]
                extra = re.match(r"\s+(\d)(?!\d)", after)
                if extra:
                    combined = day * 10 + int(extra.group(1))
                    if combined <= 31:
                        day = combined
            if 2020 <= year <= 2099 and 1 <= month <= 12 and 1 <= day <= 31:
                return year, month, day
    return None


def apply_business_rules(entry: Entry, amount: int) -> int:
    if amount <= 0:
        return 0
    if entry.account_no == 10:   # 회의식대 1만원 한도
        return min(amount, 10000)
    return amount


def run_ocr_for_entry(entry: Entry, path: Path) -> Entry:
    text = run_ocr(path)
    raw_amount = extract_amount(text)
    entry.ocr_text = text[:1000]
    entry.ocr_amount = raw_amount
    entry.amount = apply_business_rules(entry, raw_amount)
    # 파일명에 날짜가 없으면 OCR에서 추출한 날짜로 보강
    _, parsed_month, _, _ = parse_filename(Path(entry.filename).stem)
    if parsed_month is None:
        ocr_date = extract_date_from_text(text)
        if ocr_date:
            y, m, d = ocr_date
            entry.date = f"{y:04d}.{m:02d}.{d:02d}"
    return entry


# ─────────────────────────────────────────────────────────────────────────────
# Excel — openpyxl로 4월 템플릿을 그대로 쓰고 행만 채워 넣는다
# ─────────────────────────────────────────────────────────────────────────────

def write_xlsx(entries: list[dict], year: int, month: int, output_path: Path) -> None:
    wb = openpyxl.load_workbook(TEMPLATE_XLSX)
    ws = wb.worksheets[EXPENSE_SHEET_INDEX]

    ws[TITLE_CELL] = f"{year}년 {month}월분 법인(개인) 카드 정산서"

    # 기존 데이터 행 초기화 (B~H 컬럼만; J~N은 계정 분류표라 건드리지 않음)
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        for col in ("B", "C", "D", "E", "F", "G", "H"):
            ws[f"{col}{row}"] = None

    # 새 데이터 기록
    for index, entry in enumerate(entries[: DATA_END_ROW - DATA_START_ROW + 1]):
        row = DATA_START_ROW + index
        account_no = int(entry["account_no"])
        ws[f"B{row}"] = entry["date"]
        ws[f"C{row}"] = account_no
        ws[f"D{row}"] = f"=+VLOOKUP($C{row},$J$9:$L$76,3,FALSE)"
        ws[f"E{row}"] = int(entry["amount"] or 0)
        ws[f"F{row}"] = entry.get("purpose", "")
        ws[f"G{row}"] = entry.get("attendee", "정시내")
        ws[f"H{row}"] = entry.get("user", "정시내(개인)")

    # N열 계정별 합계 공식의 합산 범위를 D9:D76 / E9:E76 으로 통일 (9행도 포함)
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        ws[f"N{row}"] = f"=+SUMIF($D$9:$D$76,$L{row},$E$9:$E$76)"

    wb.save(output_path)


# ─────────────────────────────────────────────────────────────────────────────
# PDF — 영수증을 하나의 증빙 PDF로 병합
# ─────────────────────────────────────────────────────────────────────────────

def image_to_pdf_bytes(path: Path) -> bytes:
    with Image.open(path) as img:
        page = ImageOps.exif_transpose(img).convert("RGB")
        canvas = Image.new("RGB", (1240, 1754), "white")  # A4 150dpi
        page.thumbnail((1120, 1634), Image.Resampling.LANCZOS)
        x = (canvas.width - page.width) // 2
        y = (canvas.height - page.height) // 2
        canvas.paste(page, (x, y))
        buf = io.BytesIO()
        canvas.save(buf, "PDF", resolution=150)
        return buf.getvalue()


def _date_sort_key(path: Path) -> tuple:
    """파일명 `MMDD_*` 기준 날짜순 정렬 키. 파싱 실패 시 뒤로."""
    _, month, day, _ = parse_filename(path.stem)
    if month is None:
        return (99, 99, path.name)
    return (month, day, path.name)


def merge_evidence_pdf(files: list[Path], output_path: Path) -> None:
    """전달받은 파일 순서 그대로 PDF 페이지를 이어붙인다."""
    if not files:
        raise ValueError("증빙으로 만들 파일이 없습니다.")
    writer = PdfWriter()
    for src in files:
        suffix = src.suffix.lower()
        if suffix == ".pdf":
            reader = PdfReader(str(src))
        elif suffix in IMAGE_EXTS:
            reader = PdfReader(io.BytesIO(image_to_pdf_bytes(src)))
        else:
            continue
        for page in reader.pages:
            writer.add_page(page)
    with output_path.open("wb") as f:
        writer.write(f)


# ─────────────────────────────────────────────────────────────────────────────
# 세션 경로
# ─────────────────────────────────────────────────────────────────────────────

def session_paths(session_id: str) -> tuple[Path, Path]:
    root = WORK_DIR / session_id
    return root, root / "uploads"


def valid_session(session_id: str | None) -> bool:
    return bool(session_id) and bool(re.fullmatch(r"[0-9a-f]{32}", session_id))


# ─────────────────────────────────────────────────────────────────────────────
# HTML
# ─────────────────────────────────────────────────────────────────────────────

PAGE = """<!doctype html>
<html lang="ko">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>지출결의서 자동 생성 · NEOSPECTRA</title>
<style>
  /* NEOSPECTRA CI colors */
  :root {
    --neo-blue: #0070B0;
    --neo-blue-dark: #005A8C;
    --neo-blue-soft: #E6F2F9;
    --neo-green: #80B040;
    --neo-green-dark: #689632;
    --ink: #404040;
    --muted: #767676;
    --line: #E1E4E8;
    --bg: #F5F7F9;
    --white: #FFFFFF;
  }
  * { box-sizing: border-box; }
  html, body { margin:0; padding:0; }
  body {
    font-family: "Apple SD Gothic Neo", "Malgun Gothic", -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    color: var(--ink);
    background: var(--bg);
    line-height: 1.5;
    -webkit-font-smoothing: antialiased;
  }
  /* 상단 CI 액센트 바 */
  .accent-bar { height: 4px; background: linear-gradient(90deg, var(--neo-blue) 0%, var(--neo-blue) 60%, var(--neo-green) 60%, var(--neo-green) 100%); }
  header {
    background: var(--white);
    border-bottom: 1px solid var(--line);
    padding: 18px 32px;
    display: flex;
    align-items: center;
    gap: 24px;
  }
  header img.logo { height: 36px; width: auto; display: block; }
  header .divider { width: 1px; height: 28px; background: var(--line); }
  header h1 {
    margin: 0;
    font-size: 17px;
    font-weight: 600;
    color: var(--ink);
    letter-spacing: -0.2px;
  }
  header .sub { font-size: 12px; color: var(--muted); margin-top: 2px; }
  main { max-width: 1320px; margin: 0 auto; padding: 24px 32px 48px; }
  section {
    background: var(--white);
    border: 1px solid var(--line);
    border-radius: 10px;
    padding: 20px;
    margin-bottom: 16px;
    box-shadow: 0 1px 2px rgba(64,64,64,0.03);
  }
  label { display:block; font-size:12px; font-weight:500; color:var(--muted); margin-bottom:6px; letter-spacing:.2px; }
  input, select, button { font: inherit; color: inherit; }
  input, select {
    width: 100%;
    border: 1px solid var(--line);
    border-radius: 6px;
    padding: 9px 11px;
    background: var(--white);
    transition: border-color .15s, box-shadow .15s;
  }
  input:focus, select:focus {
    outline: none;
    border-color: var(--neo-blue);
    box-shadow: 0 0 0 3px rgba(0,112,176,0.12);
  }
  input[readonly] { background: #FAFBFC; color: var(--muted); }
  .grid { display:grid; grid-template-columns: 110px 110px 1fr auto; gap:12px; align-items:end; }
  .form-grid {
    display: grid;
    grid-template-columns: 100px 80px 1fr 1fr;
    gap: 12px;
    align-items: end;
  }
  .form-grid .file-field { grid-column: 1 / -1; }
  .form-grid .form-actions {
    grid-column: 1 / -1;
    display: flex;
    gap: 10px;
    justify-content: flex-end;
  }
  @media (max-width: 700px) {
    .form-grid { grid-template-columns: 1fr 1fr; }
  }
  button {
    border: 1px solid var(--neo-blue);
    border-radius: 6px;
    padding: 9px 18px;
    background: var(--neo-blue);
    color: var(--white);
    font-weight: 500;
    cursor: pointer;
    transition: background .15s;
  }
  button:hover:not(:disabled) { background: var(--neo-blue-dark); border-color: var(--neo-blue-dark); }
  button:disabled { opacity: .4; cursor: not-allowed; }
  button.secondary { background: var(--white); color: var(--neo-blue); }
  button.secondary:hover:not(:disabled) { background: var(--neo-blue-soft); color: var(--neo-blue-dark); }
  table { width:100%; border-collapse:collapse; table-layout:fixed; }
  th, td { border-bottom: 1px solid var(--line); padding: 8px 6px; text-align:left; vertical-align:middle; font-size:13px; }
  th { background: #FAFBFC; color: var(--muted); font-weight: 500; font-size: 12px; letter-spacing: .3px; text-transform: none; }
  td input, td select { padding: 7px 9px; font-size: 13px; }
  .amount { text-align:right; font-variant-numeric: tabular-nums; }
  .actions { display:flex; gap:10px; justify-content:flex-end; align-items:center; margin-bottom:12px; }
  .status { color: var(--muted); font-size:13px; flex: 1; }
  .totals {
    display:flex;
    gap: 20px;
    margin-bottom: 14px;
    font-size: 13px;
    color: var(--muted);
    justify-content: flex-end;
    align-items: baseline;
  }
  .totals strong {
    font-size: 18px;
    color: var(--neo-blue);
    font-weight: 600;
    font-variant-numeric: tabular-nums;
    margin-left: 4px;
  }
  .totals .sep { color: var(--line); }
  .downloads {
    display: flex;
    gap: 10px;
    flex-wrap: wrap;
  }
  .downloads a {
    display: inline-flex;
    align-items: center;
    gap: 8px;
    padding: 10px 16px;
    border: 1px solid var(--neo-blue);
    border-radius: 6px;
    color: var(--neo-blue);
    background: var(--white);
    text-decoration: none;
    font-weight: 500;
    font-size: 14px;
    transition: background .15s;
  }
  .downloads a:hover { background: var(--neo-blue-soft); }
  .downloads a.pdf { border-color: var(--neo-green); color: var(--neo-green-dark); }
  .downloads a.pdf:hover { background: #F1F8E8; }
  .badge {
    display: inline-block;
    padding: 2px 8px;
    border-radius: 10px;
    font-size: 11px;
    background: var(--neo-blue-soft);
    color: var(--neo-blue-dark);
    margin-left: 8px;
  }
  .table-wrap { overflow-x: auto; }
  /* OCR 실패 행 강조 */
  tr.warn td { background: #FEF6F2; }
  tr.warn td.drag-handle { box-shadow: inset 3px 0 0 #DC4A2D; color: #DC4A2D; }
  tr.warn .amount-cell { position: relative; }
  tr.warn .amount-cell::after {
    content: "확인필요";
    position: absolute;
    right: 6px;
    top: 50%;
    transform: translateY(-50%);
    font-size: 10px;
    color: #DC4A2D;
    background: #FCE6DF;
    padding: 1px 5px;
    border-radius: 8px;
    pointer-events: none;
  }
  tr.warn input[data-k="amount"] { padding-right: 60px; border-color: #F4B5A1; }
  /* 드래그 핸들 */
  td.drag-handle {
    cursor: grab;
    text-align: center;
    color: var(--muted);
    user-select: none;
    width: 36px;
  }
  td.drag-handle:hover { color: var(--neo-blue); background: var(--neo-blue-soft); }
  td.drag-handle .drag-icon { font-size: 14px; letter-spacing: -2px; line-height: 1; }
  tr.dragging { opacity: 0.4; }
  tr.drag-over { box-shadow: inset 0 -2px 0 var(--neo-blue); }
  /* 파일명 셀: 클릭 가능 표시 */
  td.file-cell { cursor: pointer; position: relative; }
  td.file-cell:hover { background: var(--neo-blue-soft); }
  td.file-cell .filename {
    display: block;
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
    color: var(--neo-blue);
    text-decoration: underline;
    text-decoration-style: dotted;
    text-underline-offset: 3px;
  }
  /* 영수증 미리보기 모달 */
  .modal-backdrop {
    position: fixed; inset: 0;
    background: rgba(20,30,40,0.6);
    display: none;
    align-items: center;
    justify-content: center;
    z-index: 50;
    padding: 24px;
  }
  .modal-backdrop.open { display: flex; }
  .modal {
    background: var(--white);
    border-radius: 10px;
    max-width: min(900px, 100%);
    max-height: 100%;
    width: 100%;
    display: flex;
    flex-direction: column;
    overflow: hidden;
    box-shadow: 0 20px 60px rgba(0,0,0,0.25);
  }
  .modal-head {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 14px 18px;
    border-bottom: 1px solid var(--line);
  }
  .modal-head h3 { margin:0; font-size: 14px; font-weight: 600; color: var(--ink); }
  .modal-close {
    border: none;
    background: transparent;
    color: var(--muted);
    cursor: pointer;
    padding: 4px 10px;
    border-radius: 4px;
    font-size: 22px;
    line-height: 1;
  }
  .modal-close:hover { background: var(--bg); color: var(--ink); }
  .modal-body {
    flex: 1;
    overflow: auto;
    background: #2A2F33;
    display: flex;
    align-items: center;
    justify-content: center;
    padding: 12px;
    min-height: 60vh;
  }
  .modal-body img { max-width: 100%; max-height: 80vh; background: #fff; box-shadow: 0 4px 12px rgba(0,0,0,0.3); }
  .modal-body iframe { width: 100%; height: 80vh; border: none; background: #fff; }
  @media (max-width: 900px) {
    main { padding: 16px; }
    header { padding: 14px 16px; gap: 12px; }
    header img.logo { height: 28px; }
    .grid { grid-template-columns: 1fr 1fr; }
    table { min-width: 1080px; }
  }
</style>
</head>
<body>
<div class="accent-bar"></div>
<header>
  <img class="logo" src="/ci/NeoSpectra_Kor_FullColor(White).png" alt="NEOSPECTRA">
  <div class="divider"></div>
  <div>
    <h1>지출결의서 자동 생성</h1>
    <div class="sub">매출전표 업로드 → OCR → 결의서 엑셀 · 증빙 PDF</div>
  </div>
</header>
<main>
  <section>
    <form id="uploadForm" class="form-grid">
      <div><label>연도</label><input name="year" type="number" value="2026" min="2020" max="2099"></div>
      <div><label>월</label><input name="month" type="number" value="5" min="1" max="12"></div>
      <div><label>참석자</label><input name="attendee" type="text" value="정시내" placeholder="예: 정시내"></div>
      <div><label>사용자</label><input name="user" type="text" value="정시내(개인)" placeholder="예: 정시내(개인)"></div>
      <div class="file-field"><label>매출전표 폴더 (이미지/PDF)</label><input name="files" type="file" webkitdirectory directory multiple accept="image/*,application/pdf"></div>
      <div class="form-actions">
        <button type="button" class="secondary" id="applyNamesBtn" title="현재 입력된 참석자/사용자를 모든 행에 다시 적용">참석자/사용자 일괄 적용</button>
        <button type="submit">불러오기</button>
      </div>
    </form>
  </section>
  <section>
    <div class="actions">
      <span id="status" class="status">매출전표 폴더를 선택하세요.</span>
      <button id="generateBtn" type="button" disabled>엑셀/PDF 생성</button>
    </div>
    <div class="totals">
      <span>OCR 합계 <strong id="ocrTotal">0</strong>원</span>
      <span>지출결의서 합계 <strong id="amountTotal">0</strong>원</span>
    </div>
    <div class="table-wrap">
      <table>
        <thead>
          <tr>
            <th style="width:36px"></th>
            <th style="width:100px">일자</th>
            <th style="width:150px">계정</th>
            <th style="width:100px">금액</th>
            <th style="width:100px">OCR</th>
            <th>목적</th>
            <th style="width:90px">참석자</th>
            <th style="width:120px">사용자</th>
            <th style="width:200px">파일</th>
          </tr>
        </thead>
        <tbody id="rows"></tbody>
      </table>
    </div>
  </section>
  <section id="downloads" class="downloads" hidden></section>
</main>
<div class="modal-backdrop" id="modal" role="dialog" aria-modal="true">
  <div class="modal">
    <div class="modal-head">
      <h3 id="modalTitle">영수증</h3>
      <button type="button" class="modal-close" id="modalClose" aria-label="닫기">×</button>
    </div>
    <div class="modal-body" id="modalBody"></div>
  </div>
</div>
<script>
const ACCOUNTS = {{ accounts|tojson }};
const ACCOUNT_GROUPS = {{ account_groups|tojson }};
let sessionId = null;

function accountOptionsHtml(selectedNo) {
  return ACCOUNT_GROUPS.map(g => {
    const opts = g.items.map(it =>
      `<option value="${it.no}" ${it.no===selectedNo?'selected':''}>${it.no} ${it.name}</option>`
    ).join('');
    return `<optgroup label="${g.category}">${opts}</optgroup>`;
  }).join('');
}

const rowsEl = document.getElementById('rows');
const statusEl = document.getElementById('status');
const generateBtn = document.getElementById('generateBtn');
const downloadsEl = document.getElementById('downloads');
const ocrTotalEl = document.getElementById('ocrTotal');
const amountTotalEl = document.getElementById('amountTotal');
const SKIP = ['지출결의서','지출증빙','거래목록','검토용','정산서'];
const IMAGE_EXT = ['.jpg','.jpeg','.png','.bmp','.webp','.tif','.tiff'];

function esc(v) { return String(v ?? '').replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c])); }

function receiptKey(file) {
  const name = file.name;
  const dot = name.lastIndexOf('.');
  if (dot < 0) return null;
  const ext = name.slice(dot).toLowerCase();
  if (![...IMAGE_EXT, '.pdf'].includes(ext)) return null;
  const stem = name.slice(0, dot);
  if (SKIP.some(m => stem.includes(m))) return null;
  const m = stem.match(/^(\\d{2})(\\d{2})[_\\-\\s]*(.+)$/);
  if (m) {
    const month = +m[1], day = +m[2];
    if (month >= 1 && month <= 12 && day >= 1 && day <= 31) {
      return `${m[1]}${m[2]}_${m[3].replace(/\\s+/g,'').toLowerCase()}`;
    }
  }
  // MMDD 없는 파일도 통과 (stem 자체를 키로)
  return stem.replace(/\\s+/g,'').toLowerCase();
}

function preferPdf(existing, next) {
  const ex = existing.name.slice(existing.name.lastIndexOf('.')).toLowerCase();
  const nx = next.name.slice(next.name.lastIndexOf('.')).toLowerCase();
  return IMAGE_EXT.includes(ex) && nx === '.pdf';
}

function filterFiles(files) {
  const map = new Map();
  for (const f of files) {
    const k = receiptKey(f);
    if (!k) continue;
    if (!map.has(k) || preferPdf(map.get(k), f)) map.set(k, f);
  }
  return [...map.values()].sort((a,b) => a.name.localeCompare(b.name, 'ko'));
}

document.getElementById('uploadForm').addEventListener('submit', async e => {
  e.preventDefault();
  const form = e.currentTarget;
  const files = filterFiles([...form.elements.files.files]);
  if (!files.length) { statusEl.textContent = '업로드 가능한 영수증 파일이 없습니다.'; return; }
  const fd = new FormData();
  fd.append('year', form.elements.year.value);
  fd.append('month', form.elements.month.value);
  files.forEach(f => fd.append('files', f, f.name));
  fd.append('attendee', form.elements.attendee.value);
  fd.append('user', form.elements.user.value);
  statusEl.textContent = `${files.length}개 영수증 업로드 중...`;
  try {
    const res = await fetch('/upload', { method:'POST', body: fd });
    const data = await res.json();
    if (!res.ok) throw new Error(data.error || '업로드 실패');
    sessionId = data.session_id;
    renderRows(data.entries);
    generateBtn.disabled = true;
    downloadsEl.hidden = true;
    statusEl.textContent = `${data.entries.length}건 로드 완료. OCR 시작...`;
    await runOcrAll();
  } catch (err) {
    statusEl.textContent = `오류: ${err.message}`;
  }
});

document.getElementById('applyNamesBtn').addEventListener('click', () => {
  const form = document.getElementById('uploadForm');
  const attendee = form.elements.attendee.value || '정시내';
  const user = form.elements.user.value || '정시내(개인)';
  const trs = [...rowsEl.querySelectorAll('tr')];
  if (!trs.length) { statusEl.textContent = '먼저 영수증을 불러오세요.'; return; }
  trs.forEach(tr => {
    tr.querySelector('[data-k="attendee"]').value = attendee;
    tr.querySelector('[data-k="user"]').value = user;
  });
  statusEl.textContent = `${trs.length}개 행의 참석자/사용자를 일괄 적용했습니다.`;
  updateTotals();
});

generateBtn.addEventListener('click', async () => {
  const entries = collectRows();
  statusEl.textContent = '엑셀/PDF 생성 중...';
  generateBtn.disabled = true;
  try {
    const res = await fetch('/generate', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({ session_id: sessionId, entries })
    });
    const data = await res.json();
    if (!res.ok) throw new Error(data.error || '생성 실패');
    statusEl.textContent = `완료 — ${entries.length}건이 처리되었습니다.`;
    downloadsEl.hidden = false;
    downloadsEl.innerHTML = `
      <a href="${data.xlsx}">
        <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/><line x1="8" y1="13" x2="16" y2="13"/><line x1="8" y1="17" x2="13" y2="17"/></svg>
        지출결의서 엑셀
      </a>
      <a class="pdf" href="${data.pdf}">
        <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/></svg>
        지출증빙 PDF
      </a>`;
  } catch (err) {
    statusEl.textContent = `오류: ${err.message}`;
  } finally {
    generateBtn.disabled = false;
  }
});

function renderRows(entries) {
  rowsEl.innerHTML = '';
  for (const e of entries) {
    const tr = document.createElement('tr');
    tr.draggable = true;
    tr.innerHTML = `
      <td class="drag-handle" title="드래그하여 순서 변경"><span class="drag-icon">⋮⋮</span></td>
      <td><input data-k="date" value="${esc(e.date)}"></td>
      <td><select data-k="account_no">${accountOptionsHtml(e.account_no)}</select></td>
      <td class="amount-cell"><input data-k="amount" class="amount" type="number" min="0" value="${e.amount || 0}"></td>
      <td><input data-k="ocr_amount" class="amount" type="number" value="${e.ocr_amount || 0}" readonly title="${esc(e.ocr_text || '')}"><input data-k="ocr_text" type="hidden" value="${esc(e.ocr_text || '')}"></td>
      <td><input data-k="purpose" value="${esc(e.purpose)}"></td>
      <td><input data-k="attendee" value="${esc(e.attendee)}"></td>
      <td><input data-k="user" value="${esc(e.user)}"></td>
      <td class="file-cell" title="클릭하여 영수증 보기"><span class="filename">${esc(e.filename)}</span><input data-k="filename" type="hidden" value="${esc(e.filename)}"><input data-k="vendor" type="hidden" value="${esc(e.vendor)}"></td>`;
    rowsEl.appendChild(tr);
  }
  rowsEl.querySelectorAll('input,select').forEach(el => el.addEventListener('input', updateTotals));
  rowsEl.querySelectorAll('td.file-cell').forEach(cell => {
    cell.addEventListener('click', () => {
      const fn = cell.querySelector('[data-k="filename"]').value;
      openPreview(fn);
    });
  });
  attachRowDragHandlers();
  updateTotals();
}

let draggingRow = null;
function attachRowDragHandlers() {
  rowsEl.querySelectorAll('tr').forEach(tr => {
    tr.addEventListener('dragstart', e => {
      draggingRow = tr;
      tr.classList.add('dragging');
      e.dataTransfer.effectAllowed = 'move';
    });
    tr.addEventListener('dragend', () => {
      tr.classList.remove('dragging');
      rowsEl.querySelectorAll('tr.drag-over').forEach(r => r.classList.remove('drag-over'));
      draggingRow = null;
      updateTotals();
    });
    tr.addEventListener('dragover', e => {
      e.preventDefault();
      if (!draggingRow || draggingRow === tr) return;
      e.dataTransfer.dropEffect = 'move';
      const rect = tr.getBoundingClientRect();
      const before = (e.clientY - rect.top) < rect.height / 2;
      rowsEl.querySelectorAll('tr.drag-over').forEach(r => r.classList.remove('drag-over'));
      tr.classList.add('drag-over');
      if (before) tr.parentNode.insertBefore(draggingRow, tr);
      else tr.parentNode.insertBefore(draggingRow, tr.nextSibling);
    });
  });
}

const modal = document.getElementById('modal');
const modalTitle = document.getElementById('modalTitle');
const modalBody = document.getElementById('modalBody');
document.getElementById('modalClose').addEventListener('click', closePreview);
modal.addEventListener('click', e => { if (e.target === modal) closePreview(); });
document.addEventListener('keydown', e => { if (e.key === 'Escape') closePreview(); });

function openPreview(filename) {
  if (!sessionId || !filename) return;
  modalTitle.textContent = filename;
  const url = `/preview/${sessionId}/${encodeURIComponent(filename)}`;
  const ext = filename.slice(filename.lastIndexOf('.')).toLowerCase();
  if (ext === '.pdf') {
    modalBody.innerHTML = `<iframe src="${url}#view=FitH" title="${esc(filename)}"></iframe>`;
  } else {
    modalBody.innerHTML = `<img src="${url}" alt="${esc(filename)}">`;
  }
  modal.classList.add('open');
}
function closePreview() {
  modal.classList.remove('open');
  modalBody.innerHTML = '';
}

function collectRows() {
  return [...rowsEl.querySelectorAll('tr')].map(tr => {
    const row = {};
    tr.querySelectorAll('[data-k]').forEach(el => row[el.dataset.k] = el.value);
    row.account_no = +row.account_no;
    row.account = ACCOUNTS[row.account_no];
    row.amount = +(row.amount || 0);
    row.ocr_amount = +(row.ocr_amount || 0);
    return row;
  });
}

async function runOcrAll() {
  const trs = [...rowsEl.querySelectorAll('tr')];
  for (let i = 0; i < trs.length; i++) {
    statusEl.textContent = `OCR ${i+1}/${trs.length}`;
    const entry = collectRows()[i];
    try {
      const res = await fetch('/ocr', {
        method:'POST', headers:{'Content-Type':'application/json'},
        body: JSON.stringify({ session_id: sessionId, entry })
      });
      const data = await res.json();
      if (!res.ok) continue;
      const tr = trs[i];
      tr.querySelector('[data-k="amount"]').value = data.entry.amount || 0;
      tr.querySelector('[data-k="ocr_amount"]').value = data.entry.ocr_amount || 0;
      tr.querySelector('[data-k="ocr_amount"]').title = data.entry.ocr_text || '';
      tr.querySelector('[data-k="ocr_text"]').value = data.entry.ocr_text || '';
      // OCR이 더 정확한 날짜를 잡았으면 반영
      if (data.entry.date) tr.querySelector('[data-k="date"]').value = data.entry.date;
      updateTotals();
    } catch (err) { /* 한 건 실패는 무시 */ }
  }
  statusEl.textContent = `OCR 완료. 금액 확인 후 '엑셀/PDF 생성'을 누르세요.`;
  generateBtn.disabled = trs.length === 0;
}

function updateTotals() {
  const rows = collectRows();
  amountTotalEl.textContent = rows.reduce((s,r)=>s+(+r.amount||0),0).toLocaleString('ko-KR');
  ocrTotalEl.textContent = rows.reduce((s,r)=>s+(+r.ocr_amount||0),0).toLocaleString('ko-KR');
  [...rowsEl.querySelectorAll('tr')].forEach((tr, i) => {
    tr.classList.toggle('warn', (+rows[i]?.amount || 0) === 0);
  });
}
</script>
</body>
</html>"""


# ─────────────────────────────────────────────────────────────────────────────
# Routes
# ─────────────────────────────────────────────────────────────────────────────

def _account_groups() -> list[dict]:
    groups: dict[str, list[dict]] = {}
    order: list[str] = []
    for no in sorted(ACCOUNTS):
        cat = ACCOUNT_CATEGORY.get(no, "기타")
        if cat not in groups:
            groups[cat] = []
            order.append(cat)
        groups[cat].append({"no": no, "name": ACCOUNTS[no]})
    return [{"category": cat, "items": groups[cat]} for cat in order]


@app.route("/")
def index():
    resp = Response(render_template_string(
        PAGE,
        accounts={str(k): v for k, v in ACCOUNTS.items()},
        account_groups=_account_groups(),
    ))
    resp.headers["Cache-Control"] = "no-store"
    return resp


@app.get("/ci/<path:filename>")
def ci_asset(filename: str):
    path = CI_DIR / clean_name(filename)
    if not path.exists() or path.suffix.lower() not in {".png", ".jpg", ".jpeg", ".svg"}:
        return Response("not found", status=404)
    return send_file(path, max_age=3600)


@app.get("/preview/<session_id>/<path:filename>")
def preview(session_id: str, filename: str):
    if not valid_session(session_id):
        return Response("bad session", status=400)
    _, upload_dir = session_paths(session_id)
    path = upload_dir / clean_name(filename)
    if not path.exists() or path.suffix.lower() not in UPLOAD_EXTS:
        return Response("not found", status=404)
    return send_file(path)


@app.post("/upload")
def upload():
    ensure_dirs()
    year = int(request.form.get("year") or datetime.now().year)
    month = int(request.form.get("month") or datetime.now().month)
    attendee = (request.form.get("attendee") or "").strip() or "정시내"
    user_name = (request.form.get("user") or "").strip() or "정시내(개인)"
    files = request.files.getlist("files")
    if not files:
        return jsonify(error="업로드할 파일이 없습니다."), 400

    session_id = uuid.uuid4().hex
    _, upload_dir = session_paths(session_id)
    upload_dir.mkdir(parents=True, exist_ok=True)

    selected: dict[str, tuple[str, object]] = {}
    for item in files:
        name = clean_name(item.filename)
        key = receipt_key(name)
        if not key:
            continue
        if key not in selected or prefer_pdf(selected[key][0], name):
            selected[key] = (name, item)

    saved: list[Path] = []
    for name, item in selected.values():
        target = upload_dir / name
        suffix = 1
        while target.exists():
            target = upload_dir / f"{Path(name).stem}_{suffix}{Path(name).suffix}"
            suffix += 1
        item.save(target)
        saved.append(target)

    saved.sort(key=lambda p: p.name)
    entries = [asdict(build_entry(p.name, year, month, attendee, user_name)) for p in saved]
    return jsonify(session_id=session_id, entries=entries)


@app.post("/ocr")
def ocr_one():
    payload = request.get_json(force=True) or {}
    session_id = payload.get("session_id")
    entry_data = payload.get("entry") or {}
    if not valid_session(session_id):
        return jsonify(error="세션 정보가 올바르지 않습니다."), 400
    filename = clean_name(entry_data.get("filename") or "")
    if not filename:
        return jsonify(error="파일명이 없습니다."), 400
    _, upload_dir = session_paths(session_id)
    path = upload_dir / filename
    if not path.exists():
        return jsonify(error=f"파일 없음: {filename}"), 404

    entry = Entry(
        date=entry_data.get("date", ""),
        account_no=int(entry_data.get("account_no") or 10),
        account=entry_data.get("account", ""),
        amount=int(entry_data.get("amount") or 0),
        purpose=entry_data.get("purpose", ""),
        attendee=entry_data.get("attendee", "정시내"),
        user=entry_data.get("user", "정시내(개인)"),
        vendor=entry_data.get("vendor", ""),
        filename=filename,
    )
    run_ocr_for_entry(entry, path)
    return jsonify(entry=asdict(entry))


@app.post("/generate")
def generate():
    ensure_dirs()
    payload = request.get_json(force=True) or {}
    session_id = payload.get("session_id")
    entries = payload.get("entries") or []
    if not valid_session(session_id):
        return jsonify(error="세션 정보가 올바르지 않습니다."), 400
    if not entries:
        return jsonify(error="기록할 거래가 없습니다."), 400
    if len(entries) > (DATA_END_ROW - DATA_START_ROW + 1):
        return jsonify(error=f"엑셀 입력 가능 행은 최대 {DATA_END_ROW - DATA_START_ROW + 1}건입니다."), 400
    if not TEMPLATE_XLSX.exists():
        return jsonify(error=f"템플릿이 없습니다: {TEMPLATE_XLSX}"), 500

    _, upload_dir = session_paths(session_id)
    if not upload_dir.exists():
        return jsonify(error="업로드 파일이 없습니다. 다시 불러오세요."), 400

    # 날짜에서 연/월 추출
    year, month = datetime.now().year, datetime.now().month
    for e in entries:
        m = re.match(r"(\d{4})\.(\d{2})\.(\d{2})", e.get("date", ""))
        if m:
            year = int(m.group(1))
            month = int(m.group(2))
            break

    out_root = OUTPUT_DIR / session_id
    out_root.mkdir(parents=True, exist_ok=True)
    yy = str(year)[2:]
    attendee_name = ""
    for e in entries:
        candidate = (e.get("attendee") or "").strip()
        if candidate:
            attendee_name = candidate
            break
    attendee_name = clean_name(attendee_name) or "정시내"
    xlsx_name = f"{yy}년_{month:02d}월_지출결의서_{attendee_name}.xlsx"
    pdf_name = f"{yy}년_{month:02d}월_지출결의서_{attendee_name}_증빙.pdf"

    write_xlsx(entries, year, month, out_root / xlsx_name)

    # PDF 병합은 사용자가 정렬한 entries 순서 그대로 따른다 (행 순서 변경 반영)
    ordered_files: list[Path] = []
    seen: set[Path] = set()
    for e in entries:
        fname = clean_name(e.get("filename") or "")
        if not fname:
            continue
        p = upload_dir / fname
        if p.exists() and p not in seen:
            ordered_files.append(p)
            seen.add(p)
    # 혹시 누락된 업로드 파일이 있으면 뒤에 추가
    for p in sorted(upload_dir.iterdir(), key=lambda x: x.name):
        if p not in seen:
            ordered_files.append(p)
    merge_evidence_pdf(ordered_files, out_root / pdf_name)

    return jsonify(
        xlsx=url_for("download", session_id=session_id, filename=xlsx_name),
        pdf=url_for("download", session_id=session_id, filename=pdf_name),
    )


@app.get("/download/<session_id>/<path:filename>")
def download(session_id: str, filename: str):
    if not valid_session(session_id):
        return Response("bad session", status=400)
    path = OUTPUT_DIR / session_id / clean_name(filename)
    if not path.exists():
        return Response("not found", status=404)
    return send_file(path, as_attachment=True, download_name=path.name)


if __name__ == "__main__":
    ensure_dirs()
    app.run(host="127.0.0.1", port=int(os.environ.get("PORT", "5000")), debug=False)
